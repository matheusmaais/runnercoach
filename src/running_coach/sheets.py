from __future__ import annotations

import csv
import json
import re
import tempfile
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


REQUIRED_SHEETS = [
    "Dashboard",
    "Estado Atual",
    "Próximos Treinos",
    "Treinos Garmin",
    "Semanas",
    "Provas e Marcos",
    "Decisões",
    "Roadmap",
    "Ciência & Prompt",
    "Dados Gráficos",
]

NAVY = "102A43"
TEAL = "14B8A6"
TEAL_DARK = "0F766E"
LIGHT_BG = "F3F6F8"
WHITE = "FFFFFF"
GRAY_TEXT = "475569"
MUTED = "E2E8F0"
AMBER = "F59E0B"
RED = "DC2626"
GREEN = "16A34A"

OPERATIONAL_SHEETS = {
    "Próximos Treinos",
    "Treinos Garmin",
    "Semanas",
    "Provas e Marcos",
    "Decisões",
    "Roadmap",
    "Ciência & Prompt",
    "Dados Gráficos",
}


@dataclass(frozen=True)
class TableData:
    headers: list[str]
    rows: list[dict[str, Any]]


def build_dashboard_workbook(
    repo_root: Path | str = Path("."),
    output_path: Path | str | None = None,
) -> Path:
    repo_root = Path(repo_root)
    output_path = Path(output_path or repo_root / "reports/dashboard.xlsx")

    data = _load_dashboard_data(repo_root)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for title in REQUIRED_SHEETS:
        workbook.create_sheet(title)

    workbook.properties.title = "Projeto Meia Forte Janeiro 2027 — Matheus & Bruna"
    workbook.properties.subject = "Running coach operational dashboard"
    workbook.properties.creator = "running-coach-system"
    workbook.properties.lastModifiedBy = "running-coach-system"
    workbook.properties.created = datetime(2026, 5, 29, 0, 0, 0)
    workbook.properties.modified = datetime(2026, 5, 29, 0, 0, 0)

    _build_dashboard(workbook["Dashboard"], data)
    _build_state_sheet(workbook["Estado Atual"], data)
    _write_table_sheet(workbook["Próximos Treinos"], data["planned"])
    _write_table_sheet(workbook["Treinos Garmin"], data["workouts"])
    _write_table_sheet(workbook["Semanas"], data["weeks"])
    _write_table_sheet(workbook["Provas e Marcos"], data["races"])
    _write_table_sheet(workbook["Decisões"], data["decisions"])
    _write_table_sheet(workbook["Roadmap"], data["roadmap"])
    _write_table_sheet(workbook["Ciência & Prompt"], data["science"])
    _write_table_sheet(workbook["Dados Gráficos"], data["chart_data"])

    for sheet_name in OPERATIONAL_SHEETS:
        _apply_operational_sheet_contract(workbook[sheet_name])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    _normalize_xlsx_zip(output_path)
    return output_path


def build_dashboard(output: Path | str) -> None:
    build_dashboard_workbook(Path(".").resolve(), Path(output))


def _load_dashboard_data(repo_root: Path) -> dict[str, TableData | dict[str, str]]:
    workouts = _read_csv_table(
        repo_root / "data/processed/workouts.csv",
        [
            "local_date",
            "activity_type",
            "category",
            "distance_km",
            "duration",
            "avg_pace",
            "bruna_pse",
            "matheus_achilles_after",
            "missing_evidence",
            "recommendation_action",
            "confidence",
        ],
    )
    planned = _read_csv_table(
        repo_root / "data/plan/planned_workouts.csv",
        [
            "planned_workout_id",
            "week_number",
            "date",
            "phase",
            "slot",
            "intended_category",
            "purpose",
            "planned_distance_or_duration",
            "planned_intensity_range",
            "allowed_fallbacks",
            "contraindications",
            "status",
        ],
    )
    decisions = _read_csv_table(
        repo_root / "data/processed/decisions.csv",
        [
            "local_date",
            "local_datetime",
            "timezone",
            "event",
            "decision",
            "reason",
            "impact",
            "related_workout_id",
            "confidence",
            "missing_evidence",
        ],
    )
    science = _read_csv_table(
        repo_root / "data/processed/science_refs.csv",
        [
            "science_ref_id",
            "title",
            "year",
            "source_type",
            "journal_or_publisher",
            "doi_or_url",
            "tags",
            "approved",
        ],
    )
    races = _read_csv_table(
        repo_root / "data/processed/races.csv",
        [
            "date",
            "race_name",
            "source",
            "status",
            "evidence_level",
            "distance_km",
            "time",
            "avg_pace",
            "bruna_avg_hr",
            "bruna_max_hr",
            "interpretation",
            "next_decision",
        ],
        placeholder={
            "date": "2026-05-31",
            "race_name": "10K balizamento",
            "source": "planned_marker",
            "status": "planned",
            "evidence_level": "none",
            "distance_km": "10",
            "time": "a definir",
            "avg_pace": "6:15-6:20",
            "bruna_avg_hr": "",
            "bruna_max_hr": "",
            "interpretation": "Prova de calibração, não treino aleatório.",
            "next_decision": "Recuperar 2-4 dias e recalibrar zonas.",
        },
    )
    roadmap = _roadmap_table()
    weeks = _weekly_summary(workouts.rows)
    chart_data = _chart_data(workouts.rows, weeks.rows)
    state = _state_summary(workouts.rows, planned.rows, decisions.rows, science.rows)

    return {
        "workouts": workouts,
        "planned": planned,
        "decisions": decisions,
        "science": science,
        "races": races,
        "roadmap": roadmap,
        "weeks": weeks,
        "chart_data": chart_data,
        "state": state,
    }


def _read_csv_table(
    path: Path,
    fallback_headers: list[str],
    placeholder: dict[str, Any] | None = None,
) -> TableData:
    if not path.exists():
        return TableData(fallback_headers, [placeholder] if placeholder else [])

    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = reader.fieldnames or fallback_headers
        rows = [dict(row) for row in reader]
    if not rows and placeholder:
        rows = [placeholder]
    return TableData(list(headers), rows)


def _roadmap_table() -> TableData:
    rows = [
        {
            "fase": "1. Polimento 10K",
            "janela": "agora",
            "objetivo": "Chegar inteiro ao 10K de balizamento.",
            "treinos_chave": "Qualidade controlada, longão leve, prova diagnóstica.",
            "riscos": "Vôlei + perna pesada + intensidade; Aquiles do Matheus.",
            "criterio": "Sem sintomas fortes e recuperação boa.",
        },
        {
            "fase": "2. Recuperação pós-10K",
            "janela": "2-7 dias pós-prova",
            "objetivo": "Absorver a prova sem compensar carga.",
            "treinos_chave": "Off/easy conforme PSE, sono e sintomas.",
            "riscos": "Tratar all-out como treino comum.",
            "criterio": "PSE normalizada, sono ok, Aquiles silencioso.",
        },
        {
            "fase": "3. Construção 5K/10K",
            "janela": "junho-agosto",
            "objetivo": "Consolidar ritmo forte controlado.",
            "treinos_chave": "Cruise intervals, strides dosados, longão leve.",
            "riscos": "Acelerar por ego ou pace instantâneo.",
            "criterio": "6:00-6:15/km controlado em blocos.",
        },
        {
            "fase": "4. Base específica meia",
            "janela": "setembro-outubro",
            "objetivo": "Aumentar tolerância aeróbia sem salto agressivo.",
            "treinos_chave": "Longões progressivos leves e controle de fadiga.",
            "riscos": "Volume extra por treino perdido.",
            "criterio": "Sem duas semanas seguidas de fadiga alta.",
        },
        {
            "fase": "5. Específico meia",
            "janela": "novembro-dezembro",
            "objetivo": "Converter 5K/10K em meia forte.",
            "treinos_chave": "Blocos sustentados e longões com controle.",
            "riscos": "Empilhar intensidade e força pesada.",
            "criterio": "Longões consistentes e recuperação preservada.",
        },
        {
            "fase": "6. Polimento meia",
            "janela": "janeiro 2027",
            "objetivo": "Chegar fresco para correr forte.",
            "treinos_chave": "Redução de volume, estímulos curtos controlados.",
            "riscos": "Testar limite perto da prova.",
            "criterio": "Saúde > heroísmo.",
        },
    ]
    return TableData(list(rows[0]), rows)


def _weekly_summary(workouts: list[dict[str, Any]]) -> TableData:
    weeks: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "semana": "",
            "volume_km": 0.0,
            "treinos": 0,
            "qualidade": 0,
            "longoes": 0,
            "pse_max": "",
            "risco": "baixo",
        }
    )
    for row in workouts:
        parsed_date = _parse_date(row.get("local_date") or row.get("date"))
        week_key = _week_key(parsed_date)
        bucket = weeks[week_key]
        bucket["semana"] = week_key
        bucket["volume_km"] += _to_float(row.get("distance_km"))
        bucket["treinos"] += 1
        category = str(row.get("category") or "").lower()
        if any(token in category for token in ["quality", "cruise", "race", "threshold"]):
            bucket["qualidade"] += 1
        if _to_float(row.get("distance_km")) >= 8:
            bucket["longoes"] += 1
        pse = _to_int(row.get("bruna_pse"))
        if pse is not None:
            bucket["pse_max"] = max(_to_int(bucket["pse_max"]) or 0, pse)
        if _risk_score(row) >= 70:
            bucket["risco"] = "alto"
        elif _risk_score(row) >= 40 and bucket["risco"] != "alto":
            bucket["risco"] = "atenção"

    rows = sorted(weeks.values(), key=lambda item: item["semana"])
    for row in rows:
        row["volume_km"] = round(row["volume_km"], 2)
    return TableData(
        ["semana", "volume_km", "treinos", "qualidade", "longoes", "pse_max", "risco"],
        rows,
    )


def _chart_data(workouts: list[dict[str, Any]], weeks: list[dict[str, Any]]) -> TableData:
    rows: list[dict[str, Any]] = []
    weekly_volume = {row["semana"]: row["volume_km"] for row in weeks}
    for index, row in enumerate(sorted(workouts, key=lambda item: item.get("local_date", "")), 1):
        parsed_date = _parse_date(row.get("local_date") or row.get("date"))
        week = _week_key(parsed_date)
        pace_seconds = _pace_to_seconds(row.get("avg_pace"))
        rows.append(
            {
                "ordem": index,
                "data": row.get("local_date") or row.get("date") or "",
                "semana": week,
                "distancia_km": _to_float(row.get("distance_km")),
                "pace_min_km": round(pace_seconds / 60, 2) if pace_seconds else "",
                "volume_semana_km": weekly_volume.get(week, 0),
                "risco_score": _risk_score(row),
            }
        )
    if not rows:
        rows.append(
            {
                "ordem": 1,
                "data": str(date.today()),
                "semana": "sem dados",
                "distancia_km": 0,
                "pace_min_km": "",
                "volume_semana_km": 0,
                "risco_score": 0,
            }
        )
    return TableData(list(rows[0]), rows)


def _state_summary(
    workouts: list[dict[str, Any]],
    planned: list[dict[str, Any]],
    decisions: list[dict[str, Any]],
    science: list[dict[str, Any]],
) -> dict[str, str]:
    latest_shared = max(
        (row for row in workouts if _is_shared_coaching_evidence(row)),
        key=_workout_sort_key,
        default={},
    )
    latest_solo = max(
        (
            row
            for row in workouts
            if row.get("athlete_context") == "matheus_garmin_only"
        ),
        key=_workout_sort_key,
        default={},
    )
    next_workout = next(
        (
            row
            for row in sorted(planned, key=lambda item: item.get("date", "9999-99-99"))
            if row.get("status") == "planned"
        ),
        {},
    )
    avg_recent_pace = latest_shared.get("avg_pace")
    risk = _risk_label(latest_shared)
    tiredness = _tiredness_label(latest_shared)
    approved_science = sum(str(row.get("approved")).lower() == "true" for row in science)
    decision = decisions[-1].get("decision", "sem decisão") if decisions else "sem decisão"
    matheus_solo_note = (
        f" Último Garmin solo do Matheus: {latest_solo.get('avg_pace')} "
        "(não usado para evolução da Bruna)."
        if latest_solo.get("avg_pace")
        else ""
    )
    evolucao = (
        f"Último treino compartilhado com check-in: pace médio {avg_recent_pace}; "
        "tendência ainda pede mais check-ins."
        if avg_recent_pace
        else "Sem evidência compartilhada suficiente; não inferir evolução da Bruna."
    )

    return {
        "evolucao": f"{evolucao}{matheus_solo_note}",
        "cansaco": tiredness,
        "risco": risk,
        "proximo_treino": _format_next_workout(next_workout),
        "meia_forte": "No caminho se longão seguir leve, qualidade for controlada e decisões continuarem auditáveis.",
        "ultima_decisao": decision,
        "fontes_aprovadas": str(approved_science),
    }


def _build_dashboard(ws, data: dict[str, TableData | dict[str, str]]) -> None:
    state = data["state"]
    assert isinstance(state, dict)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A1"
    ws.merge_cells("A1:H2")
    title = ws["A1"]
    title.value = "Projeto Meia Forte Janeiro 2027 — Matheus & Bruna"
    title.font = Font(color=WHITE, bold=True, size=18)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    cards = [
        ("Evolução", state["evolucao"], GREEN),
        ("Cansaço", state["cansaco"], AMBER),
        ("Risco", state["risco"], RED if "alto" in state["risco"].lower() else TEAL),
        ("Próximo treino", state["proximo_treino"], TEAL),
        ("Meia forte", state["meia_forte"], NAVY),
    ]
    positions = ["A4:B7", "C4:D7", "E4:F7", "G4:H7", "A9:H12"]
    for (label, value, accent), cell_range in zip(cards, positions, strict=True):
        _dashboard_card(ws, cell_range, label, value, accent)

    ws["A14"] = "Tendência operacional"
    ws["A14"].font = Font(bold=True, color=NAVY, size=14)
    ws["A15"] = "Fonte: aba Dados Gráficos; pace em min/km, risco em escala 0-100."
    ws["A15"].font = Font(color=GRAY_TEXT, italic=True)

    _add_dashboard_chart(ws)

    ws["F14"] = "Controles"
    ws["F14"].font = Font(bold=True, color=NAVY, size=14)
    controls = [
        ("Decisão atual", state["ultima_decisao"]),
        ("Fontes aprovadas", state["fontes_aprovadas"]),
        ("Regra central", "Consistência > heroísmo"),
        ("Status", "Dashboard reexecutável"),
    ]
    for offset, (label, value) in enumerate(controls, start=15):
        ws[f"F{offset}"] = label
        ws[f"G{offset}"] = value
        ws[f"F{offset}"].font = Font(bold=True, color=GRAY_TEXT)
        ws[f"G{offset}"].alignment = Alignment(wrap_text=True)

    for column, width in {
        "A": 16,
        "B": 18,
        "C": 16,
        "D": 18,
        "E": 16,
        "F": 18,
        "G": 20,
        "H": 20,
    }.items():
        ws.column_dimensions[column].width = width
    for row in range(1, 24):
        ws.row_dimensions[row].height = 24


def _dashboard_card(ws, cell_range: str, label: str, value: str, accent: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = f"{label}\n{value}"
    cell.fill = PatternFill("solid", fgColor=WHITE)
    cell.font = Font(color=NAVY, bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border_side = Side(style="thin", color=MUTED)
    cell.border = Border(
        left=Side(style="thick", color=accent),
        right=border_side,
        top=border_side,
        bottom=border_side,
    )


def _add_dashboard_chart(ws) -> None:
    chart = LineChart()
    chart.title = "Pace médio e risco por treino"
    chart.style = 13
    chart.y_axis.title = "min/km ou risco"
    chart.x_axis.title = "Treinos"
    chart.height = 7.2
    chart.width = 14
    data = Reference(ws.parent["Dados Gráficos"], min_col=5, max_col=7, min_row=1, max_row=30)
    cats = Reference(ws.parent["Dados Gráficos"], min_col=2, min_row=2, max_row=30)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend.position = "b"
    ws.add_chart(chart, "A17")


def _build_state_sheet(ws, data: dict[str, TableData | dict[str, str]]) -> None:
    state = data["state"]
    assert isinstance(state, dict)
    rows = [
        {"item": "Evolução", "estado": state["evolucao"]},
        {"item": "Cansaço", "estado": state["cansaco"]},
        {"item": "Risco", "estado": state["risco"]},
        {"item": "Próximo treino", "estado": state["proximo_treino"]},
        {"item": "Caminho meia forte", "estado": state["meia_forte"]},
        {"item": "Fontes científicas aprovadas", "estado": state["fontes_aprovadas"]},
    ]
    _write_table_sheet(ws, TableData(["item", "estado"], rows))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_table_sheet(ws, table: TableData) -> None:
    ws.sheet_view.showGridLines = False
    headers = table.headers or ["status"]
    rows = table.rows or [{headers[0]: "sem dados"}]

    for col_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(rows, 2):
        for col_index, header in enumerate(headers, 1):
            cell = ws.cell(row=row_index, column=col_index, value=_display_value(row.get(header)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_BG)

    last_row = max(2, len(rows) + 1)
    last_col = max(1, len(headers))
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    if ws.title in OPERATIONAL_SHEETS:
        excel_table = Table(displayName=_safe_table_name(ws.title), ref=table_ref)
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(excel_table)
    _size_columns(ws, headers, rows)


def _apply_operational_sheet_contract(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _size_columns(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    for col_index, header in enumerate(headers, 1):
        values = [header] + [str(row.get(header) or "") for row in rows[:50]]
        max_len = max(len(value) for value in values)
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 12), 42)


def _display_value(value: Any) -> Any:
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, bool):
        return "true" if value else "false"
    return value


def _safe_table_name(sheet_name: str) -> str:
    cleaned = "".join(char for char in sheet_name if char.isalnum())
    return f"tbl_{cleaned[:24]}"


def _format_next_workout(row: dict[str, Any]) -> str:
    if not row:
        return "Sem próximo treino planejado."
    return (
        f"{row.get('date', '')} · {row.get('purpose', '')} · "
        f"{row.get('planned_distance_or_duration', '')} · "
        f"{row.get('planned_intensity_range', '')}"
    )


def _risk_label(row: dict[str, Any]) -> str:
    if not row:
        return "incerto: sem check-in compartilhado suficiente."
    score = _risk_score(row)
    if score >= 70:
        return f"alto ({score}/100): reduzir e resolver evidência."
    if score >= 40:
        return f"atenção ({score}/100): manter conservador."
    return f"baixo ({score}/100): seguir sem heroísmo."


def _tiredness_label(row: dict[str, Any]) -> str:
    if not row:
        return "incerto: falta check-in compartilhado."
    pse = _to_int(row.get("bruna_pse"))
    missing = _json_list(row.get("missing_evidence"))
    if pse is not None and pse >= 9:
        return "alto: PSE >= 9 pede leve/off."
    if "checkin" in missing:
        return "incerto: falta check-in subjetivo."
    if row.get("volleyball_previous_day") == "true":
        return "atenção: vôlei no dia anterior conta carga."
    return "controlado pelos dados disponíveis."


def _risk_score(row: dict[str, Any]) -> int:
    score = 15
    pse = _to_int(row.get("bruna_pse"))
    achilles = _to_int(row.get("matheus_achilles_after"))
    missing = _json_list(row.get("missing_evidence"))
    if pse is not None:
        score += max(0, pse - 6) * 12
    if achilles is not None:
        score += max(0, achilles - 2) * 15
    if row.get("volleyball_previous_day") == "true":
        score += 15
    if "checkin" in missing:
        score += 20
    return min(score, 100)


def _is_shared_coaching_evidence(row: dict[str, Any]) -> bool:
    if row.get("athlete_context") != "shared_run_with_manual_checkin":
        return False
    if str(row.get("shared_run")).lower() != "true":
        return False
    if str(row.get("bruna_present")).lower() != "true":
        return False
    return "checkin" not in _json_list(row.get("missing_evidence"))


def _workout_sort_key(row: dict[str, Any]) -> str:
    return str(row.get("local_datetime") or row.get("local_date") or row.get("date") or "")


def _parse_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _week_key(value: date | None) -> str:
    if value is None:
        return "sem-data"
    iso = value.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _pace_to_seconds(value: Any) -> int | None:
    if not value:
        return None
    parts = str(value).split(":")
    if len(parts) != 2:
        return None
    try:
        return int(parts[0]) * 60 + int(parts[1])
    except ValueError:
        return None


def _json_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    if not value:
        return []
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list):
        return []
    return [str(item) for item in parsed]


def _normalize_xlsx_zip(path: Path) -> None:
    fixed_date = (2026, 5, 29, 0, 0, 0)
    fixed_timestamp = b"2026-05-29T00:00:00Z"
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as handle:
        temp_path = Path(handle.name)

    try:
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
            temp_path, "w", compression=zipfile.ZIP_DEFLATED
        ) as target:
            for name in sorted(source.namelist()):
                info = zipfile.ZipInfo(name, date_time=fixed_date)
                info.compress_type = zipfile.ZIP_DEFLATED
                info.external_attr = source.getinfo(name).external_attr
                payload = source.read(name)
                if name == "docProps/core.xml":
                    payload = re.sub(
                        rb"<dcterms:created[^>]*>[^<]+</dcterms:created>",
                        rb'<dcterms:created xsi:type="dcterms:W3CDTF">'
                        + fixed_timestamp
                        + rb"</dcterms:created>",
                        payload,
                    )
                    payload = re.sub(
                        rb"<dcterms:modified[^>]*>[^<]+</dcterms:modified>",
                        rb'<dcterms:modified xsi:type="dcterms:W3CDTF">'
                        + fixed_timestamp
                        + rb"</dcterms:modified>",
                        payload,
                    )
                target.writestr(info, payload)
        temp_path.replace(path)
    finally:
        if temp_path.exists():
            temp_path.unlink()
