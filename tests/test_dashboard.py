from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from running_coach.sheets import REQUIRED_SHEETS, build_dashboard_workbook


def test_dashboard_workbook_contract(tmp_path):
    repo_root = tmp_path / "repo"
    _write_fixture_csvs(repo_root)

    output = build_dashboard_workbook(repo_root)

    assert output == repo_root / "reports/dashboard.xlsx"
    assert output.exists()

    workbook = load_workbook(output)
    assert workbook.sheetnames == REQUIRED_SHEETS

    dashboard = workbook["Dashboard"]
    assert dashboard["A1"].value == "Projeto Meia Forte Janeiro 2027 — Matheus & Bruna"
    assert dashboard["A4"].value.startswith("Evolução")
    assert "treino compartilhado" in dashboard["A4"].value
    assert "6:47" in dashboard["A4"].value
    assert "4:22" in dashboard["A4"].value
    assert "não usado para evolução da Bruna" in dashboard["A4"].value
    assert "local_date" not in str(dashboard["A1"].value)
    assert dashboard._charts, "Dashboard must contain at least one native chart"
    assert dashboard["A1"].fill.fgColor.rgb.endswith("102A43")
    assert dashboard["A4"].border.left.color.rgb.endswith("16A34A")

    for sheet_name in [
        "Próximos Treinos",
        "Treinos Garmin",
        "Semanas",
        "Provas e Marcos",
        "Decisões",
        "Roadmap",
        "Ciência & Prompt",
        "Dados Gráficos",
    ]:
        sheet = workbook[sheet_name]
        assert sheet.freeze_panes == "A2", sheet_name
        assert sheet.auto_filter.ref is not None, sheet_name

    assert workbook["Treinos Garmin"]["A1"].value == "local_date"
    assert workbook["Dados Gráficos"]["E1"].value == "pace_min_km"
    assert workbook["Estado Atual"]["A2"].value == "Evolução"
    assert workbook["Provas e Marcos"]["C1"].value == "source"
    assert workbook["Provas e Marcos"]["C2"].value == "planned_marker"


def test_dashboard_build_is_byte_deterministic(tmp_path):
    repo_root = tmp_path / "repo"
    _write_fixture_csvs(repo_root)

    output = build_dashboard_workbook(repo_root)
    first_hash = hashlib.sha256(output.read_bytes()).hexdigest()
    build_dashboard_workbook(repo_root)
    second_hash = hashlib.sha256(output.read_bytes()).hexdigest()

    assert second_hash == first_hash
    workbook = load_workbook(output)
    assert workbook["Dashboard"]._charts


def _write_fixture_csvs(repo_root: Path) -> None:
    processed = repo_root / "data/processed"
    plan = repo_root / "data/plan"
    processed.mkdir(parents=True)
    plan.mkdir(parents=True)

    (processed / "workouts.csv").write_text(
        "local_date,local_datetime,athlete_context,shared_run,bruna_present,"
        "activity_type,category,distance_km,duration,avg_pace,bruna_pse,"
        "matheus_achilles_after,volleyball_previous_day,missing_evidence,"
        "recommendation_action,confidence\n"
        "2026-05-28,2026-05-28 16:17:36,shared_run_with_manual_checkin,true,true,"
        'Corrida,cruise_intervals,7.47,3039,6:47,7,0,true,'
        '["bruna_avg_hr"],request_manual_resolution,low\n'
        "2026-05-28,2026-05-28 17:14:17,matheus_garmin_only,false,false,"
        'Corrida,,1.33,349,4:22,,0,false,'
        '["checkin"],request_manual_resolution,low\n',
        encoding="utf-8",
    )
    (processed / "decisions.csv").write_text(
        "local_date,local_datetime,timezone,event,decision,reason,impact,"
        "related_workout_id,confidence,missing_evidence\n"
        "2026-05-28,2026-05-28 16:17:36,America/Sao_Paulo,"
        "pipeline_after_workout,defer,Manual review,Preserve audit trail,"
        "workout-1,low,[]\n",
        encoding="utf-8",
    )
    (processed / "science_refs.csv").write_text(
        "science_ref_id,title,year,source_type,journal_or_publisher,doi_or_url,tags,approved\n"
        "threshold-1,Threshold evidence,2024,peer_reviewed_study,Journal,"
        'https://doi.org/10.1000/example,["threshold"],true\n',
        encoding="utf-8",
    )
    (plan / "planned_workouts.csv").write_text(
        "planned_workout_id,week_number,date,phase,slot,intended_category,purpose,"
        "planned_distance_or_duration,planned_intensity_range,allowed_fallbacks,"
        "contraindications,status\n"
        "plan-1,1,2026-05-31,ten_k_polish,sunday,diagnostic_race_10k,"
        "10K diagnostic race,10 km,6:15-6:20/km,"
        '["replace_with_easy"],["red_flag"],planned\n',
        encoding="utf-8",
    )
